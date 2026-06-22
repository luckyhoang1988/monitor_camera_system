"""Xuất báo cáo uptime ra file Excel (.xlsx) bằng openpyxl.

Nhận `UptimeReport` (đã áp bộ lọc khoảng thời gian + khu vực ở report_service) và
dựng workbook 2 sheet: uptime từng NVR và camera mất tín hiệu nhiều nhất.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.services.report_service import UptimeReport

_HEADER_FILL = PatternFill("solid", fgColor="0D6EFD")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13)


def _localtime(dt) -> str:
    """Định dạng datetime sang giờ local (giống filter localtime ở web)."""
    if dt is None:
        return "—"
    tz = ZoneInfo(get_settings().timezone)
    return dt.astimezone(tz).strftime("%d/%m/%Y %H:%M:%S")


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, widths: list[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_report_xlsx(report: UptimeReport, *, area: str | None = None) -> bytes:
    """Dựng workbook từ báo cáo, trả về bytes của file .xlsx."""
    wb = Workbook()

    # --- Sheet 1: Uptime theo NVR ---
    ws = wb.active
    ws.title = "Uptime NVR"
    scope = f"Khu vực: {area}" if area else "Tất cả khu vực"
    if report.is_custom_range:
        tz = ZoneInfo(get_settings().timezone)
        d_from = report.start.astimezone(tz).strftime("%d/%m/%Y") if report.start else "—"
        d_to = report.end.astimezone(tz).strftime("%d/%m/%Y") if report.end else "—"
        period = f"từ {d_from} đến {d_to}"
    else:
        period = f"{report.days} ngày gần nhất"
    ws["A1"] = f"Báo cáo uptime — {period} · {scope}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = (
        f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Uptime NVR: {report.system_nvr_uptime}%  ·  "
        f"Uptime Camera: {report.system_camera_uptime}%"
    )

    headers = ["NVR", "Khu vực", "Số lần kiểm tra", "Số lần Online", "Uptime (%)"]
    header_row = 4
    ws.append([])  # row 3 trống
    ws.append(headers)
    _style_header(ws, header_row, len(headers))
    for r in report.nvr_rows:
        ws.append(
            [r.name, r.area or "—", r.total_checks, r.online_checks, r.uptime_pct]
        )
    _autosize(ws, [34, 16, 16, 16, 12])
    ws.freeze_panes = f"A{header_row + 1}"

    # --- Sheet 2: Camera mất tín hiệu nhiều nhất ---
    ws2 = wb.create_sheet("Camera mất tín hiệu")
    headers2 = [
        "NVR", "Kênh", "Tên camera", "Số lần Offline", "Tổng kiểm tra", "Uptime (%)"
    ]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    for c in report.worst_cameras:
        ws2.append(
            [
                c.nvr_name,
                c.channel_no,
                c.name or "—",
                c.offline_checks,
                c.total_checks,
                c.uptime_pct,
            ]
        )
    _autosize(ws2, [34, 8, 28, 16, 16, 12])
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Nhật ký NVR online trở lại ---
    ws3 = wb.create_sheet("NVR online tro lai")
    headers3 = ["Thời điểm online lại", "NVR", "Khu vực", "Trạng thái trước"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    for e in report.nvr_recoveries:
        ws3.append(
            [
                _localtime(e.recovered_at),
                e.name,
                e.area or "—",
                e.from_status or "—",
            ]
        )
    _autosize(ws3, [22, 34, 16, 18])
    ws3.freeze_panes = "A2"

    # --- Sheet 4: Nhật ký camera online trở lại ---
    ws4 = wb.create_sheet("Camera online tro lai")
    headers4 = [
        "Thời điểm online lại",
        "NVR",
        "Khu vực",
        "Kênh",
        "Tên camera",
        "Trạng thái trước",
    ]
    ws4.append(headers4)
    _style_header(ws4, 1, len(headers4))
    for e in report.camera_recoveries:
        ws4.append(
            [
                _localtime(e.recovered_at),
                e.nvr_name,
                e.area or "—",
                e.channel_no,
                e.name or "—",
                e.from_status or "—",
            ]
        )
    _autosize(ws4, [22, 30, 16, 8, 28, 18])
    ws4.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_offline_cameras_xlsx(rows: list[dict]) -> bytes:
    """Xuất danh sách camera đang mất tín hiệu (bảng trên trang Tổng quan).

    `rows` là output của query_service.list_offline_cameras: mỗi phần tử có
    `camera` (CameraChannel), `nvr_name`, `nvr_area`.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Camera mất tín hiệu"

    ws["A1"] = "Danh sách camera đang mất tín hiệu"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = (
        f"Xuất lúc: {_localtime(datetime.now(ZoneInfo(get_settings().timezone)))}"
        f"  ·  Tổng: {len(rows)} camera"
    )

    headers = [
        "Đầu ghi (NVR)", "Khu vực", "Kênh", "Tên camera", "IP",
        "Trạng thái", "Mất tín hiệu từ", "Kiểm tra cuối",
    ]
    header_row = 4
    ws.append([])  # row 3 trống
    ws.append(headers)
    _style_header(ws, header_row, len(headers))

    for row in rows:
        cam = row["camera"]
        ws.append(
            [
                row["nvr_name"],
                row.get("nvr_area") or "—",
                cam.channel_no,
                cam.name or "—",
                cam.camera_ip or "—",
                cam.current_status,
                _localtime(cam.offline_since),
                _localtime(cam.last_checked_at),
            ]
        )
    _autosize(ws, [22, 14, 8, 26, 16, 12, 20, 20])
    ws.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
